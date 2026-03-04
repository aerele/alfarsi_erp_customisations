import frappe
from openpyxl import Workbook
from io import BytesIO
import json

def send_sales_person_yearly_report():
    report = frappe.get_doc("Report", "SalesPerson Yearly Report")
    columns, data = report.get_data(as_dict=True)

    subject = "Yearly Sales Person Report"
    message = "Please find the attached yearly sales person report."
    pre_data = frappe.db.get_single_value("Sales Person Yearly Report Settings", "data")
    json_pre_data = json.loads(pre_data) if pre_data else {}
    wb = Workbook()
    ws = wb.active

    for col_idx, col in enumerate(columns, 1):
        ws.cell(row=1, column=col_idx, value=col['label'])
    for row_idx, row in enumerate(data, 2):
        for col_idx, col in enumerate(columns, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col['fieldname'], '0'))
    file_name = f"Sales_Person_Yearly_Report_{frappe.utils.nowdate()}.xlsx"
    
    if json_pre_data:
        ws.cell(row=1, column=len(columns)+2, value="Previous Total")
        ws.cell(row=1, column=len(columns)+3, value="Difference")
        for row_idx, row in enumerate(data, 2):
            salesperson = row.get('salesperson')
            pre_total = row.get('total_sales')
            prev_value = json_pre_data.get(salesperson, 0)
            ws.cell(row=row_idx, column=len(columns)+2, value=prev_value)
            ws.cell(row=row_idx, column=len(columns)+3, value=pre_total - prev_value)


    file_bytes = BytesIO()
    wb.save(file_bytes)
    file_bytes.seek(0)

    file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": file_name,
            "is_private": 0,
            "content": file_bytes.read()
        })
    file_doc.save()

    attachments = [{"file_url": file_doc.file_url}]
    recipients = frappe.db.get_single_value("Sales Person Yearly Report Settings", "send_report_to")
    to_list = [email.strip() for email in recipients.split(",") if email.strip()]
    frappe.sendmail(
        recipients=to_list,
        subject=subject,
        message=message,
        attachments=attachments,
    )
    to_save = {}
    for d in data:
        to_save[d.salesperson] = d.total_sales
    frappe.db.set_single_value("Sales Person Yearly Report Settings", "data", json.dumps(to_save))