import os
import frappe
from openpyxl import Workbook
from io import BytesIO
from frappe.utils import nowdate, getdate, add_days, add_months, get_last_day, today

@frappe.whitelist()
def send_scheduled_sellout_mails():
    today = getdate(nowdate())
    settings = frappe.get_single("Brand Sellout Mail Settings")

    if not settings.mail_configuration:
        return {"success": False}

    for config_row in settings.mail_configuration:
        if not config_row.last_send_date:
            should_send = True
        else:
            next_send_date = add_days(config_row.last_send_date, config_row.frequency)
            should_send = today >= getdate(next_send_date)

        if not should_send:
            continue
        if not config_row.last_send_date:
            start_date = add_days(today, -config_row.frequency)
        else:
            start_date = getdate(config_row.last_send_date)

        end_date = today
        to_list = [mail.strip() for mail in config_row.sent_mail_to.split(",") if mail.strip()]
        cc_list = [mail.strip() for mail in config_row.cc_to.split(",") if mail.strip()]

        filters = {
            "from_date": start_date,
            "to_date": end_date,
            "brand": config_row.brand,
            "include_all_items": "Need all items from the selected brand"
        }

        report = frappe.get_doc("Report", "Sellout Report")
        columns, report_rows = report.get_data(filters=filters)
        wb = Workbook()
        ws = wb.active

        for col_idx, col in enumerate(columns, 1):
            header = col.get('label', '')
            if col.get('fieldname') == "item_code":
                header = "Item Code"
            elif col.get('fieldname') == "item_name":
                header = "Item Name"
            ws.cell(row=1, column=col_idx, value=header)

        for row_idx, data_row in enumerate(report_rows, 2):
            if isinstance(data_row, dict):
                for col_idx, col in enumerate(columns, 1):
                    value = data_row.get(col['fieldname'], "")
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)
                    ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                for col_idx, value in enumerate(data_row, 1):
                    if isinstance(value, float) and value.is_integer():
                        value = int(value)
                    ws.cell(row=row_idx, column=col_idx, value=value)

        file_name = f"brand_sellout_{config_row.brand}_{today.strftime('%Y%m%d')}.xlsx"
        file_bytes = BytesIO()
        wb.save(file_bytes)
        file_bytes.seek(0)

        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": file_name,
            "is_private": 0,
            "content": file_bytes.read()
        })
        file_doc.save()
        subject = "Sellout Report"
        message = f"""
        Hi Team,

        Please find the attached sellout data for the period:
        {start_date} → {end_date}
        """
        frappe.sendmail(
            recipients=to_list,
            cc=cc_list,
            subject=subject,
            message=message,
            attachments=[{"file_url": file_doc.file_url}],
        )
        config_row.last_send_date = today

    settings.save(ignore_permissions=True)
    frappe.db.commit()

    return {"success": True}
